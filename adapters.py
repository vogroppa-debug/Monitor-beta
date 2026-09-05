# -*- coding: utf-8 -*-
"""
adapters.py — Normaliza cada CSV tidy de indicadores CESS al ESQUEMA CANÓNICO largo.

Esquema canónico de salida (una fila = una observación):
    <dimensiones...> , metrica , valor
donde entre las dimensiones siempre hay un eje temporal (`anio` y/o `periodo`) y,
cuando aplica, una geografía (`departamento` / `unidad_geografica`).

Cada adapter devuelve un dict:
    {
      "fields":  [nombres de columna, la última siempre "valor", la anterior "metrica"],
      "rows":    [[...], ...],                 # valores; `valor` en formato inglés (decimal .)
      "dims":    {dim: [valores ordenados]},   # valores disponibles por dimensión (para filtros)
      "metricas":{clave: {"label":..., "unidad":...}},
      "notas":   [incidencias detectadas para BUILD_NOTES.md],
    }

Por defecto se lee el CSV ya generado (contrato tidy). El build no re-descarga nada.
"""
import os
import math
import unicodedata
import pandas as pd
import openpyxl

# Carpeta padre = donde viven los CSV de los indicadores.
DATA_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _src(name):
    return os.path.join(DATA_DIR, name)


def _drv(name):
    """Archivos descargados del Drive a la subcarpeta datos-drive/."""
    return os.path.join(DATA_DIR, "datos-drive", name)


def _ipc_anual():
    """IPC NOA promedio anual (base dic-2016=100). Clave: año (int)."""
    df = pd.read_csv(_src("ipc_noa_mensual.csv"), encoding="utf-8")
    df["anio"] = df["indice_tiempo"].astype(str).str[:4].astype(int)
    return df.groupby("anio")["ipc_ng_noa"].mean().to_dict()


def _xlsx_rows(path, sheet=None, header_row=0):
    """Lee una hoja xlsx (valores cacheados) -> (encabezado, [dict por fila])."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = list(data[header_row])
    out = [dict(zip(hdr, row)) for row in data[header_row + 1:]]
    return hdr, out


# --------------------------------------------------------------------------
# Normalización de nombres de departamento de Salta (Title Case con acentos).
# Los CSV vienen con distinto casing (MAYÚSCULAS en educación, Title en el resto).
# --------------------------------------------------------------------------
_DEPT_CANON = {
    "ANTA": "Anta", "CACHI": "Cachi", "CAFAYATE": "Cafayate", "CAPITAL": "Capital",
    "CERRILLOS": "Cerrillos", "CHICOANA": "Chicoana", "GENERAL GUEMES": "General Güemes",
    "GENERAL JOSE DE SAN MARTIN": "General José de San Martín", "GUACHIPAS": "Guachipas",
    "IRUYA": "Iruya", "LA CALDERA": "La Caldera", "LA CANDELARIA": "La Candelaria",
    "LA POMA": "La Poma", "LA VINA": "La Viña", "LOS ANDES": "Los Andes", "METAN": "Metán",
    "MOLINOS": "Molinos", "ORAN": "Orán", "RIVADAVIA": "Rivadavia",
    "ROSARIO DE LA FRONTERA": "Rosario de la Frontera", "ROSARIO DE LERMA": "Rosario de Lerma",
    "SAN CARLOS": "San Carlos", "SANTA VICTORIA": "Santa Victoria",
}
# Valores que NO son un departamento real (datos enmascarados / sin asignar).
NON_GEO = {"Enmascarado", "Sin datos", "Sin asignar", "Sin Asignar", "SIN DATOS", "ENMASCARADO",
           "Otros Deptos./Local. No Publicables", "Otros Deptos./Local. No Publicable"}
# Tokens de agregado provincial (no son un departamento/municipio; se excluyen de rankings).
# "Total Salta" se usa donde "Salta" ya es un municipio real (Ciudad de Salta), p. ej. construcción.
PROVINCIAL_TOKENS = {"Salta", "Total Salta"}


def _unacc_upper(s):
    return "".join(c for c in unicodedata.normalize("NFKD", str(s))
                   if not unicodedata.combining(c)).upper().strip()


def norm_dept(s):
    """Devuelve el nombre canónico del departamento, o el token especial tal cual."""
    if s is None or (isinstance(s, float) and math.isnan(s)):
        return None
    raw = str(s).strip()
    if raw in NON_GEO:
        return raw
    key = _unacc_upper(raw)
    if key in NON_GEO:
        return "Sin datos" if "DATO" in key else ("Sin asignar" if "ASIGNAR" in key else raw.title())
    return _DEPT_CANON.get(key, raw.title())


def _sorted_unique(series):
    return sorted([v for v in series.dropna().unique().tolist()], key=lambda x: str(x))


# ==========================================================================
# TEMA 1 — Educación por departamento
# ==========================================================================
EDU_METRICAS = {
    "matricula":         {"label": "Matrícula",             "unidad": "alumnos"},
    "egresados":         {"label": "Egresados",             "unidad": "alumnos"},
    "repitentes":        {"label": "Repitentes",            "unidad": "alumnos"},
    "sobreedad_alumnos": {"label": "Alumnos con sobreedad", "unidad": "alumnos"},
    "extranjeros":       {"label": "Alumnos extranjeros",   "unidad": "alumnos"},
    "cargos":            {"label": "Cargos docentes",       "unidad": "cargos"},
    "horas":             {"label": "Horas cátedra",         "unidad": "horas"},
}


def educacion():
    df = pd.read_csv(_src("educacion_departamentos_salta_tidy.csv"), encoding="utf-8-sig")
    df["departamento"] = df["departamento"].map(norm_dept)
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    notas = []
    n_bad = int(df["valor"].isna().sum())
    if n_bad:
        notas.append(f"educacion: {n_bad} valores no numéricos descartados.")
    df = df.dropna(subset=["valor"])
    fields = ["anio", "departamento", "sector", "ambito", "nivel", "metrica", "valor"]
    df = df[fields]
    return {
        "fields": fields,
        "rows": df.values.tolist(),
        "dims": {
            "anio": _sorted_unique(df["anio"]),
            "departamento": [d for d in _sorted_unique(df["departamento"]) if d not in NON_GEO],
            "sector": _sorted_unique(df["sector"]),
            "ambito": _sorted_unique(df["ambito"]),
            "nivel": ["inicial", "primaria", "secundaria"],
        },
        "metricas": EDU_METRICAS,
        "notas": notas,
    }


# ==========================================================================
# TEMA 2 — Vitivinicultura (INV)
# ==========================================================================
INV_METRICAS = {
    "produccion_uva":           {"label": "Producción de uva",        "unidad": "quintales"},
    "elaboracion_total":        {"label": "Elaboración (vino+mosto)",  "unidad": "hectolitros"},
    "despacho_mercado_interno": {"label": "Despacho al mercado interno", "unidad": "hectolitros"},
    "export_volumen":           {"label": "Exportaciones (volumen)",   "unidad": "hectolitros"},
    "export_valor_fob":         {"label": "Exportaciones (valor FOB)", "unidad": "miles US$"},
    "import_volumen":           {"label": "Importaciones (nacional)",  "unidad": "hectolitros"},
}


def inv():
    df = pd.read_csv(_src("inv_salta_tidy.csv"), encoding="utf-8-sig")
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    df = df.dropna(subset=["valor"])
    fields = ["anio", "categoria", "nivel", "unidad_geografica", "metrica", "valor"]
    df = df[fields]
    return {
        "fields": fields,
        "rows": df.values.tolist(),
        "dims": {
            "anio": _sorted_unique(df["anio"]),
            "categoria": _sorted_unique(df["categoria"]),
            "nivel": _sorted_unique(df["nivel"]),
            "unidad_geografica": _sorted_unique(df["unidad_geografica"]),
        },
        "metricas": INV_METRICAS,
        "notas": [],
    }


# ==========================================================================
# TEMA 3 — Producción de petróleo y gas (melt de medidas → metrica/valor)
# ==========================================================================
PROD_MEAS = ["prod_pet", "prod_gas", "prod_agua", "iny_agua", "iny_otro", "tef", "pozos"]
PROD_METRICAS = {
    "prod_pet":  {"label": "Producción de petróleo", "unidad": "m³"},
    "prod_gas":  {"label": "Producción de gas",      "unidad": "miles de m³"},
    "prod_agua": {"label": "Producción de agua",     "unidad": "m³"},
    "iny_agua":  {"label": "Inyección de agua",      "unidad": "m³"},
    "iny_otro":  {"label": "Inyección (otros)",      "unidad": "m³"},
    "tef":       {"label": "Tiempo efectivo",        "unidad": "pozos-mes"},
    "pozos":     {"label": "Pozos",                  "unidad": "pozos"},
}


def _melt_prod(df, id_cols):
    df = df.copy()
    df["departamento"] = df["departamento"].map(norm_dept)
    keep = id_cols + [m for m in PROD_MEAS if m in df.columns]
    long = df[keep].melt(id_vars=id_cols, var_name="metrica", value_name="valor")
    long["valor"] = pd.to_numeric(long["valor"], errors="coerce")
    long = long.dropna(subset=["valor"])
    return long


def produccion_energia():
    """Combina la serie MENSUAL (para el eje temporal) y la ANUAL (para ranking)."""
    mens = pd.read_csv(_src("Produccion_Petroleo_Gas_Salta_por_departamento_desde_2018.csv"),
                       encoding="utf-8-sig")
    anual = pd.read_csv(_src("Produccion_Petroleo_Gas_Salta_por_departamento_anual.csv"),
                        encoding="utf-8-sig")

    long_m = _melt_prod(mens, ["departamento", "anio", "mes", "tipo_de_recurso"])
    long_m["anio"] = long_m["anio"].astype(int)
    long_m["mes"] = long_m["mes"].astype(int)
    long_m["trimestre"] = [_trim_label(a, m) for a, m in zip(long_m["anio"], long_m["mes"])]

    # Descartar meses finales "stub" (dato preliminar que cae a ~0): distorsionarían el
    # trimestre en curso. Regla: un mes final cuya producción de gas es < 10 % del mes previo.
    _gm = (long_m[long_m["metrica"] == "prod_gas"].groupby(["anio", "mes"])["valor"].sum()
           .reset_index().sort_values(["anio", "mes"]))
    _v = _gm["valor"].tolist()
    _ym = list(zip(_gm["anio"].astype(int), _gm["mes"].astype(int)))
    _drop = set()
    for i in range(len(_v) - 1, 0, -1):
        if _v[i] < 0.10 * _v[i - 1]:
            _drop.add(_ym[i])
        else:
            break
    if _drop:
        long_m = long_m[~long_m.apply(lambda x: (int(x["anio"]), int(x["mes"])) in _drop, axis=1)]

    # TRIMESTRAL por tipo de recurso: los flujos (m³) se SUMAN dentro del trimestre.
    trim = (long_m.groupby(["departamento", "anio", "trimestre", "tipo_de_recurso", "metrica"],
                           as_index=False)["valor"].sum())
    trim.insert(0, "grano", "trimestral")

    # ANUAL oficial (para ranking y KPIs), tipo agregado "TOTAL".
    long_a = _melt_prod(anual, ["departamento", "anio"])
    long_a["anio"] = long_a["anio"].astype(int)
    long_a["trimestre"] = ""
    long_a["tipo_de_recurso"] = "TOTAL"
    long_a.insert(0, "grano", "anual")

    fields = ["grano", "departamento", "anio", "trimestre", "tipo_de_recurso", "metrica", "valor"]
    both = pd.concat([trim[fields], long_a[fields]], ignore_index=True)

    # El eje temporal trimestral no debe mostrar el último trimestre parcial (caída ficticia).
    trims_ok, _anios_ok, _tl = _completos(long_m, mes_col="mes")
    rows = _keep_completos(both.values.tolist(), trims_ok, anios_ok=None, gi=0, ai=2, ti=3)

    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": _sorted_unique(both["anio"]),
            "departamento": [d for d in _sorted_unique(long_a["departamento"]) if d not in NON_GEO],
            "tipo_de_recurso": ["CONVENCIONAL", "NO CONVENCIONAL", "NO DISCRIMINADO"],
            "trimestre": sorted(trims_ok, key=_trim_key),
        },
        "metricas": PROD_METRICAS,
        "notas": ["produccion_energia: se descartan iny_gas, iny_co2 y vida_util (todos 0 en Salta).",
                  "produccion_energia: serie temporal agregada por TRIMESTRE (suma de meses); "
                  "el ranking y los KPIs usan el total anual oficial."],
    }


# ==========================================================================
# TEMA 4 — Empleo registrado y remuneraciones (OEDE) por rubro y departamento
#   · empleo:         puestos registrados del sector privado (SUMABLE)
#   · remun_real_idx: índice de remuneración media real, deflactada por IPC NOA,
#                     base dic-2023 = 100, por serie (NO sumable).
# ==========================================================================
EMPLEO_METRICAS = {
    "empleo":         {"label": "Empleo registrado",                        "unidad": "puestos", "agg": "sum"},
    "remun_real_idx": {"label": "Remuneración real (índice dic-2023 = 100)", "unidad": "índice",  "agg": "mean"},
}

# Nombres de sector (rubro) del archivo OEDE: se corrigen tildes/typos de origen.
_RUBRO_LABEL = {
    "Agricultura, ganaderia y pesca":  "Agricultura, ganadería y pesca",
    "Comercio":                        "Comercio",
    "Construccion":                    "Construcción",
    "Electircidad, gas y agua":        "Electricidad, gas y agua",
    "Explotacion de minas y canteras": "Explotación de minas y canteras",
    "Industria manufacturera":         "Industria manufacturera",
    "Servicios":                       "Servicios",
    "Sin rama":                        "Sin rama",
}


def _ipc_noa_map():
    """Serie IPC Nivel General región NOA (base dic-2016=100). Clave: periodo AAAAMM (int)."""
    df = pd.read_csv(_src("ipc_noa_mensual.csv"), encoding="utf-8")
    ym = df["indice_tiempo"].astype(str).str.slice(0, 7).str.replace("-", "", regex=False).astype(int)
    return dict(zip(ym, pd.to_numeric(df["ipc_ng_noa"], errors="coerce")))


def _wide_long(name, timecol):
    """Archivo ancho (timecol × columnas de departamento) -> largo [timecol, departamento, valor]."""
    df = pd.read_csv(_src(name), encoding="utf-8-sig")
    long = df.melt(id_vars=[timecol], var_name="departamento", value_name="valor")
    long["departamento"] = long["departamento"].map(norm_dept)
    long["valor"] = pd.to_numeric(long["valor"], errors="coerce")
    return long.dropna(subset=["valor"])


def _trim_label(anio, mes):
    return f"{int(anio)}-T{(int(mes) - 1) // 3 + 1}"


def _trim_key(t):
    y, q = t.split("-T")
    return (int(y), int(q))


def _completos(mensual, mes_col="mes"):
    """A partir de un df mensual (con columnas `anio`, `trimestre` y una de mes),
    devuelve (set trimestres completos>=3 meses, set años completos>=12 meses)
    y la lista ordenada de trimestres completos. Sirve para no emitir períodos
    parciales (típicamente el último), que dibujarían una caída ficticia."""
    mc = mensual.groupby("trimestre")[mes_col].nunique()
    trims_ok = sorted([t for t, n in mc.items() if n >= 3], key=_trim_key)
    yc = mensual.groupby("anio")[mes_col].nunique()
    anios_ok = {int(a) for a, n in yc.items() if n >= 12}
    return set(trims_ok), anios_ok, trims_ok


def _keep_completos(rows, trims_ok, anios_ok=None, gi=0, ai=1, ti=2):
    """Filtra filas de series temporales: descarta trimestres incompletos y,
    si se pasa anios_ok, también años incompletos. `rows` es lista de listas con
    grano en gi, anio en ai y etiqueta de trimestre en ti."""
    out = []
    for r in rows:
        if r[gi] == "trimestral" and r[ti] not in trims_ok:
            continue
        if anios_ok is not None and r[gi] == "anual" and int(r[ai]) not in anios_ok:
            continue
        out.append(r)
    return out


def _add_tiempo(df):
    """Agrega columnas anio/mes/trimestre a un largo con `periodo` AAAAMM (int)."""
    df = df.copy()
    df["anio"] = (df["periodo"] // 100).astype(int)
    df["mes"] = (df["periodo"] % 100).astype(int)
    df["trimestre"] = [_trim_label(a, m) for a, m in zip(df["anio"], df["mes"])]
    return df


def _emit_periodos(rows, key_cols, val_col, grano_metrica, group):
    """Agrega `group` (df con anio/trimestre) por trimestre y por año con el agregador
    dado y agrega filas [grano, departamento, anio, trimestre, rubro, metrica, valor].
    grano_metrica: (rubro, metrica, agg) donde agg ∈ {'mean','sum'}."""
    rubro, metrica, agg = grano_metrica
    # Trimestral
    gq = group.groupby(key_cols + ["anio", "trimestre"], as_index=False)[val_col].agg(agg)
    for r in gq.itertuples(index=False):
        d = getattr(r, "departamento")
        rows.append(["trimestral", d, int(r.anio), getattr(r, "trimestre"),
                     rubro, metrica, round(float(getattr(r, val_col)), 1)])
    # Anual
    ga = group.groupby(key_cols + ["anio"], as_index=False)[val_col].agg(agg)
    for r in ga.itertuples(index=False):
        d = getattr(r, "departamento")
        rows.append(["anual", d, int(r.anio), "",
                     rubro, metrica, round(float(getattr(r, val_col)), 1)])


def empleo():
    ipc = _ipc_noa_map()
    BASE = 202312  # base del índice real: diciembre de 2023

    # ---- Series mensuales anchas (empleo y salario por departamento) --------
    emp_m = _add_tiempo(_wide_long("OEDE_Empleo_Salta_departamento_mensual.csv", "periodo"))
    rem_m = _add_tiempo(_wide_long("OEDE_Remuneraciones_Salta_departamento_mensual.csv", "periodo"))

    # Trimestres COMPLETOS (3 meses) -> el KPI compara trimestres completos.
    tmp = emp_m[["trimestre", "periodo"]].drop_duplicates()
    mcount = tmp.groupby("trimestre")["periodo"].nunique()
    trims_completos = sorted([t for t, n in mcount.items() if n >= 3], key=_trim_key)

    rows = []

    # ---- Empleo por RUBRO (sector) y departamento (media de meses) ----------
    det = pd.read_csv(_src("OEDE_Salta_detalle_apertura_mensual.csv"), encoding="utf-8-sig")
    sec = det[det["apertura"] == "sector"].copy()
    sec["departamento"] = sec["departamento"].map(norm_dept)
    sec["rubro"] = sec["categoria"].astype(str).map(lambda c: _RUBRO_LABEL.get(c, c))
    sec["empleo"] = pd.to_numeric(sec["empleo"], errors="coerce")
    sec = sec.dropna(subset=["empleo"])
    sec["anio"] = sec["anio"].astype(int)
    sec["trimestre"] = [_trim_label(a, m) for a, m in zip(sec["anio"], sec["mes"])]
    # media de los meses del período, por (depto, rubro)
    gq = sec.groupby(["departamento", "rubro", "anio", "trimestre"], as_index=False)["empleo"].mean()
    for r in gq.itertuples(index=False):
        rows.append(["trimestral", r.departamento, int(r.anio), r.trimestre, r.rubro, "empleo", round(float(r.empleo), 1)])
    ga = sec.groupby(["departamento", "rubro", "anio"], as_index=False)["empleo"].mean()
    for r in ga.itertuples(index=False):
        rows.append(["anual", r.departamento, int(r.anio), "", r.rubro, "empleo", round(float(r.empleo), 1)])

    # ---- Empleo TOTAL por departamento (media de meses) + provincial -------
    _emit_periodos(rows, ["departamento"], "valor", ("Total", "empleo", "mean"), emp_m)
    # Provincial = suma de los niveles departamentales (por trimestre y por año).
    emp_q = emp_m.groupby(["departamento", "anio", "trimestre"], as_index=False)["valor"].mean()
    for r in emp_q.groupby(["anio", "trimestre"], as_index=False)["valor"].sum().itertuples(index=False):
        rows.append(["trimestral", "Salta", int(r.anio), r.trimestre, "Total", "empleo", round(float(r.valor), 1)])
    emp_y = emp_m.groupby(["departamento", "anio"], as_index=False)["valor"].mean()
    for r in emp_y.groupby("anio", as_index=False)["valor"].sum().itertuples(index=False):
        rows.append(["anual", "Salta", int(r.anio), "", "Total", "empleo", round(float(r.valor), 1)])

    # ---- Remuneración real (índice base dic-2023=100), media de meses ------
    rem_m["ipc"] = rem_m["periodo"].map(ipc)
    rem_m = rem_m.dropna(subset=["ipc"])
    rem_m["real"] = rem_m["valor"] / rem_m["ipc"]
    base_dep = rem_m[rem_m["periodo"] == BASE].set_index("departamento")["real"].to_dict()
    rem_m["idx"] = [100.0 * real / base_dep[dep] if base_dep.get(dep) else None
                    for real, dep in zip(rem_m["real"], rem_m["departamento"])]
    rem_m = rem_m.dropna(subset=["idx"])
    _emit_periodos(rows, ["departamento"], "idx", ("Total", "remun_real_idx", "mean"), rem_m)

    # Provincial: media ponderada por empleo -> masa salarial / empleo total.
    mm = emp_m.merge(rem_m[["periodo", "departamento", "valor", "ipc"]], on=["periodo", "departamento"],
                     suffixes=("_emp", "_sal"))
    mm["mass"] = mm["valor_emp"] * mm["valor_sal"]
    prov = mm.groupby("periodo", as_index=False).agg(emp=("valor_emp", "sum"), mass=("mass", "sum"), ipc=("ipc", "first"))
    prov["salprov"] = prov["mass"] / prov["emp"]
    prov["real"] = prov["salprov"] / prov["ipc"]
    base_prov = prov.loc[prov["periodo"] == BASE, "real"]
    base_prov = float(base_prov.iloc[0]) if len(base_prov) else None
    if base_prov:
        prov["idx"] = 100.0 * prov["real"] / base_prov
        prov = _add_tiempo(prov).dropna(subset=["idx"])
        for r in prov.groupby(["anio", "trimestre"], as_index=False)["idx"].mean().itertuples(index=False):
            rows.append(["trimestral", "Salta", int(r.anio), r.trimestre, "Total", "remun_real_idx", round(float(r.idx), 1)])
        for r in prov.groupby("anio", as_index=False)["idx"].mean().itertuples(index=False):
            rows.append(["anual", "Salta", int(r.anio), "", "Total", "remun_real_idx", round(float(r.idx), 1)])

    fields = ["grano", "departamento", "anio", "trimestre", "rubro", "metrica", "valor"]
    deptos = [d for d in sorted({r[1] for r in rows}) if d not in NON_GEO and d not in PROVINCIAL_TOKENS]
    rubros = sorted({r[4] for r in rows if r[4] != "Total"})
    trimestres = sorted({r[3] for r in rows if r[3]}, key=_trim_key)
    anios = sorted({int(r[2]) for r in rows})
    notas = [
        "empleo: remuneración expresada como ÍNDICE REAL (deflactado por IPC NOA, "
        "serie 145.3_INGNOANOA_DICI_M_10 de datos.gob.ar), base dic-2023 = 100 por serie.",
        "empleo: agregación TRIMESTRAL y ANUAL (media de los meses del período para empleo e "
        "índice real). Se incluyen períodos parciales (p. ej. 2025); el KPI compara el último "
        "trimestre completo contra el mismo trimestre del año anterior.",
        "empleo: la remuneración real provincial es media ponderada por empleo "
        "(Σ empleo·salario / Σ empleo); no se suman promedios entre rubros ni departamentos.",
    ]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": anios,
            "departamento": deptos,
            "rubro": rubros + ["Total"],
            "trimestre": trimestres,
        },
        "metricas": EMPLEO_METRICAS,
        "trims_completos": trims_completos,
        "notas": notas,
    }


# ==========================================================================
# TEMA — Turismo (EOH INDEC), consolidado provincial Salta
# ==========================================================================
TURISMO_METRICAS = {
    "viajeros":         {"label": "Viajeros",       "unidad": "viajeros",         "agg": "sum"},
    "pernoctaciones":   {"label": "Pernoctaciones",  "unidad": "noches",           "agg": "sum"},
    "ocupacion_hab":    {"label": "Ocupación de habitaciones", "unidad": "%",       "agg": "mean"},
    "estadia":          {"label": "Estadía media",   "unidad": "noches",           "agg": "mean"},
    "establecimientos": {"label": "Establecimientos", "unidad": "establecimientos", "agg": "mean"},
}


def turismo():
    df = pd.read_csv(_drv("Salta_turismo_EOH_mensual.csv"), encoding="utf-8-sig")
    df = df[(df["ambito"] == "localidad") & (df["provincia"] == "Salta")].copy()
    num = ["habitaciones_disponibles", "habitaciones_ocupadas", "viajeros", "viajeros_residentes",
           "viajeros_no_residentes", "pernoctaciones", "pernoctaciones_residentes",
           "pernoctaciones_no_residentes", "establecimientos"]
    for c in num:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["anio"] = df["anio"].astype(int)
    df["n_mes"] = df["n_mes"].astype(int)
    df["trimestre"] = [_trim_label(a, m) for a, m in zip(df["anio"], df["n_mes"])]
    df["mes_lbl"] = [f"{a}-{m:02d}" for a, m in zip(df["anio"], df["n_mes"])]

    # Consolidar las localidades de Salta por mes (sumar conteos).
    mens = df.groupby(["anio", "n_mes", "mes_lbl", "trimestre"], as_index=False)[num].sum()

    rows = []
    # segmentos de viajeros y pernoctaciones
    SEG_V = [("Total", "viajeros"), ("Residentes", "viajeros_residentes"), ("No residentes", "viajeros_no_residentes")]
    SEG_P = [("Total", "pernoctaciones"), ("Residentes", "pernoctaciones_residentes"), ("No residentes", "pernoctaciones_no_residentes")]

    def push(grano, mes_lbl, trim, anio, sub):
        # sub: dict con columnas ya agregadas para el período (sumas de conteos)
        hd, ho = sub["habitaciones_disponibles"], sub["habitaciones_ocupadas"]
        for seg, col in SEG_V:
            rows.append([grano, anio, mes_lbl, trim, seg, "viajeros", round(sub[col], 1)])
        for seg, col in SEG_P:
            rows.append([grano, anio, mes_lbl, trim, seg, "pernoctaciones", round(sub[col], 1)])
        if hd:
            rows.append([grano, anio, mes_lbl, trim, "Total", "ocupacion_hab", round(100.0 * ho / hd, 1)])
        if sub["viajeros"]:
            rows.append([grano, anio, mes_lbl, trim, "Total", "estadia", round(sub["pernoctaciones"] / sub["viajeros"], 2)])
        rows.append([grano, anio, mes_lbl, trim, "Total", "establecimientos", round(sub["establecimientos"], 1)])

    # Mensual (cada mes) — conteos por mes; establecimientos es stock del mes.
    for r in mens.itertuples(index=False):
        d = r._asdict()
        push("mensual", d["mes_lbl"], d["trimestre"], int(d["anio"]), d)
    # Trimestral y anual: sumar conteos de los meses; establecimientos = promedio del período.
    def agg_group(keys, grano, mes_lbl_fn, trim_fn):
        g = mens.groupby(keys, as_index=False).agg(
            {**{c: "sum" for c in num if c != "establecimientos"}, "establecimientos": "mean", "anio": "first"})
        for r in g.itertuples(index=False):
            d = r._asdict()
            push(grano, mes_lbl_fn(d), trim_fn(d), int(d["anio"]), d)
    agg_group(["anio", "trimestre"], "trimestral", lambda d: "", lambda d: d["trimestre"])
    agg_group(["anio"], "anual", lambda d: "", lambda d: "")

    # No emitir trimestres/años incompletos (la EOH termina a mitad de año => caída ficticia).
    trims_ok, anios_ok, trims_completos = _completos(mens, mes_col="n_mes")
    rows = _keep_completos(rows, trims_ok, anios_ok, gi=0, ai=1, ti=3)

    fields = ["grano", "anio", "mes", "trimestre", "segmento", "metrica", "valor"]
    anios = sorted({r[1] for r in rows})
    trimestres = sorted({r[3] for r in rows if r[3]}, key=_trim_key)
    meses = sorted({r[2] for r in rows if r[2]})
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": anios, "trimestre": trimestres, "mes": meses,
            "segmento": ["Total", "Residentes", "No residentes"],
        },
        "metricas": TURISMO_METRICAS,
        "trims_completos": trims_completos,
        "notas": [
            "turismo: consolidado provincial de Salta (Ciudad de Salta + Cafayate) a partir de la "
            "EOH-INDEC; conteos sumados entre localidades y tasas (ocupación, estadía) recalculadas.",
            "turismo: agregación trimestral/anual por suma de los meses (conteos); la EOH termina "
            "en noviembre de 2025 (serie discontinuada).",
        ],
    }


# ==========================================================================
# TEMA — Agricultura (MAGyP), por cultivo y departamento
# ==========================================================================
AGRI_METRICAS = {
    "superficie_sembrada_ha":  {"label": "Superficie sembrada",  "unidad": "ha",         "agg": "sum"},
    "superficie_cosechada_ha": {"label": "Superficie cosechada", "unidad": "ha",         "agg": "sum"},
    "produccion_tm":           {"label": "Producción",           "unidad": "toneladas",  "agg": "sum"},
    "rendimiento_kgxha":       {"label": "Rendimiento",          "unidad": "kg/ha",      "agg": "mean"},
}


def agricultura():
    wb = openpyxl.load_workbook(_drv("CESS_Agricultura.xlsx"), read_only=True, data_only=True)
    frames = []
    for sh in wb.sheetnames:
        if sh.endswith("1a") or sh.endswith("2a"):   # evita doble conteo con "Soja" total
            continue
        data = list(wb[sh].iter_rows(values_only=True))
        if not data:
            continue
        hdr = list(data[0])
        frames.append(pd.DataFrame(data[1:], columns=hdr))
    wb.close()
    df = pd.concat(frames, ignore_index=True)
    df["departamento"] = df["departamento"].map(norm_dept)
    df["cultivo"] = df["cultivo"].astype(str).str.capitalize()
    for c in ["anio", "superficie_sembrada_ha", "superficie_cosechada_ha", "produccion_tm", "rendimiento_kgxha"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(subset=["anio"])
    df["anio"] = df["anio"].astype(int)

    # Descartar la(s) campaña(s) final(es) claramente parciales (aún sin los cultivos de
    # verano): si el total producido de un año es < 50% del anterior, no está completo.
    tot = df.groupby("anio")["produccion_tm"].sum().sort_index()
    years = list(tot.index)
    while len(years) >= 2 and tot[years[-1]] < 0.5 * tot[years[-2]]:
        years.pop()
    df = df[df["anio"].isin(years)]

    rows = []
    metcols = ["superficie_sembrada_ha", "superficie_cosechada_ha", "produccion_tm"]
    # Nivel departamento (valores tal cual; rendimiento de origen).
    for r in df.itertuples(index=False):
        d = r._asdict()
        for m in metcols:
            if pd.notna(d[m]):
                rows.append([int(d["anio"]), d["departamento"], d["cultivo"], m, round(float(d[m]), 1)])
        if pd.notna(d["rendimiento_kgxha"]):
            rows.append([int(d["anio"]), d["departamento"], d["cultivo"], "rendimiento_kgxha", round(float(d["rendimiento_kgxha"]), 1)])
    # Provincial "Salta" por (anio, cultivo): sumar superficies/producción; rendimiento ponderado.
    prov = df.groupby(["anio", "cultivo"], as_index=False)[metcols].sum()
    for r in prov.itertuples(index=False):
        d = r._asdict()
        for m in metcols:
            rows.append([int(d["anio"]), "Salta", d["cultivo"], m, round(float(d[m]), 1)])
        sc = d["superficie_cosechada_ha"]
        if sc:
            rows.append([int(d["anio"]), "Salta", d["cultivo"], "rendimiento_kgxha",
                         round(1000.0 * d["produccion_tm"] / sc, 1)])

    fields = ["anio", "departamento", "cultivo", "metrica", "valor"]
    deptos = [d for d in sorted({r[1] for r in rows}) if d not in NON_GEO and d not in PROVINCIAL_TOKENS]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({r[0] for r in rows}),
            "departamento": deptos,
            "cultivo": sorted({r[2] for r in rows}),
        },
        "metricas": AGRI_METRICAS,
        "notas": [
            "agricultura: campañas 2018/19–2025/26 (MAGyP); el año indica el inicio de campaña.",
            "agricultura: 'Salta' es el total provincial (suma de departamentos); el rendimiento "
            "provincial se calcula como producción/superficie cosechada, no como promedio simple.",
        ],
    }


# ==========================================================================
# TEMA — Gobierno / Finanzas públicas provinciales (ejecución del gasto)
# ==========================================================================
GOB_METRICAS = {
    "gasto_corr": {"label": "Gasto (pesos corrientes)",  "unidad": "pesos", "agg": "sum"},
    "gasto_real": {"label": "Gasto (pesos constantes)",  "unidad": "pesos", "agg": "sum"},
}
_GOB_BASE_Y = 2025  # los reales se expresan en pesos del último año ejecutado
_GOB_PERSONAL = "Gastos en personal"
# Etiquetas legibles por código de objeto (el concepto del PDF viene en mayúsculas y a veces truncado).
_OBJ_LABELS = {
    "1": "Gastos en personal", "2": "Bienes de consumo", "3": "Servicios no personales",
    "4": "Bienes de uso", "5": "Transferencias", "6": "Activos financieros",
    "7": "Servicio de la deuda", "8": "Otros gastos", "9": "Gastos figurativos",
}


def gobierno():
    ipc = _ipc_anual()
    base = ipc[_GOB_BASE_Y]

    def real(v, y):
        return v * base / ipc[y] if ipc.get(y) else None

    rows = []
    YEARS = [2021, 2022, 2023, 2024, 2025]

    # ---- Objeto del gasto (ejecutado = compromisos) ------------------------
    _, obj = _xlsx_rows(_src("Ejecucion_Compromisos_Objeto_Gasto_Salta_2021-2025.xlsx"), header_row=4)
    tot = {y: 0.0 for y in YEARS}
    for r in obj:
        code = str(r.get("Código") or "").strip()
        concepto = r.get("Concepto")
        if not code or code in ("None",) or concepto is None:
            continue
        if len(code) != 1 or not code.isdigit():   # solo los rubros de primer nivel
            continue
        partida = _OBJ_LABELS.get(code, str(concepto).strip().capitalize())
        for y in YEARS:
            v = r.get(str(y))
            if v is None:
                continue
            v = float(v)
            rows.append([y, "objeto", "principal", partida, "gasto_corr", round(v, 1)])
            rows.append([y, "objeto", "principal", partida, "gasto_real", round(real(v, y), 1)])
            tot[y] += v
    for y in YEARS:
        rows.append([y, "objeto", "total", "Total ejecutado", "gasto_corr", round(tot[y], 1)])
        rows.append([y, "objeto", "total", "Total ejecutado", "gasto_real", round(real(tot[y], y), 1)])

    # ---- Finalidad (hoja Serie (total)) ------------------------------------
    _, fin = _xlsx_rows(_src("Ejecucion_Finalidad_Funcion_Salta_2021-2025.xlsx"),
                        sheet="Serie (total)", header_row=4)
    for r in fin:
        code = str(r.get("Código") or "").strip()
        nombre = r.get("Finalidad / Función / Subfunción")
        if not code or "." in code or not code.isdigit() or nombre is None:
            continue  # solo finalidades (código de un dígito, sin punto)
        partida = str(nombre).strip().title()
        for y in YEARS:
            v = r.get(str(y))
            if v is None:
                continue
            v = float(v)
            rows.append([y, "finalidad", "finalidad", partida, "gasto_corr", round(v, 1)])
            rows.append([y, "finalidad", "finalidad", partida, "gasto_real", round(real(v, y), 1)])

    # ---- Transferencias corrientes vs capital (del tidy) -------------------
    tr = pd.read_csv(_src("transferencias_ff_salta_tidy.csv"), encoding="utf-8-sig")
    tr = tr[tr["code"].astype(str) == "0"]
    for r in tr.itertuples(index=False):
        y = int(r.anio)
        if y not in YEARS:
            continue
        for etq, col in [("Corrientes", "corrientes"), ("Capital", "capital")]:
            v = float(getattr(r, col))
            rows.append([y, "transferencia", "tipo", etq, "gasto_corr", round(v, 1)])
            rows.append([y, "transferencia", "tipo", etq, "gasto_real", round(real(v, y), 1)])

    fields = ["anio", "clasificador", "nivel", "partida", "metrica", "valor"]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": YEARS,
            "clasificador": ["objeto", "finalidad", "transferencia"],
            "nivel": ["principal", "total", "finalidad", "tipo"],
            "partida": sorted({r[3] for r in rows}),
        },
        "metricas": GOB_METRICAS,
        "notas": [
            "gobierno: ejecución del gasto provincial (consolidado Adm. Central + Organismos "
            "Descentralizados), acumulada a diciembre; fuente presupuesto.salta.gob.ar.",
            "gobierno: 'objeto' usa los compromisos ejecutados; los pesos constantes se deflactan "
            "por el IPC NOA (promedio anual), base 2025.",
        ],
    }


# ==========================================================================
# TEMA — Ganadería bovina (MAGyP), stock por departamento y categoría
# ==========================================================================
GAN_METRICAS = {
    "stock_bovino":  {"label": "Stock bovino",         "unidad": "cabezas", "agg": "sum"},
    "unidades_prod": {"label": "Unidades productivas", "unidad": "UP",      "agg": "sum"},
}
_GAN_CAT = {
    "total_bovinos": "Total", "vacas": "Vacas", "vaquillonas": "Vaquillonas",
    "novillos": "Novillos", "novillitos": "Novillitos", "terneros": "Terneros",
    "terneras": "Terneras", "toros": "Toros", "toritos": "Toritos", "bueyes": "Bueyes",
}


def ganaderia():
    df = pd.read_csv(_src("Ganaderia_Stock_Bovino_Salta_departamento_2012-2025.csv"), encoding="utf-8-sig")
    df["departamento"] = df["departamento"].map(norm_dept)
    df["anio"] = pd.to_numeric(df["anio"], errors="coerce").astype("Int64")
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
    df = df.dropna(subset=["anio", "valor"])
    df["anio"] = df["anio"].astype(int)

    rows = []
    for r in df.itertuples(index=False):
        cat = _GAN_CAT.get(r.metrica)
        if cat is not None:
            rows.append([int(r.anio), r.departamento, cat, "stock_bovino", round(float(r.valor), 1)])
        elif r.metrica == "cantidad_up":
            rows.append([int(r.anio), r.departamento, "Total", "unidades_prod", round(float(r.valor), 1)])
    # Provincial "Salta" = suma de departamentos por (anio, categoria, metrica).
    prov = {}
    for a, dep, cat, met, val in rows:
        prov[(a, cat, met)] = prov.get((a, cat, met), 0.0) + val
    prov_rows = [[a, "Salta", cat, met, round(v, 1)] for (a, cat, met), v in prov.items()]

    allrows = rows + prov_rows
    fields = ["anio", "departamento", "categoria", "metrica", "valor"]
    deptos = [d for d in sorted({r[1] for r in allrows}) if d not in NON_GEO and d not in PROVINCIAL_TOKENS]
    return {
        "fields": fields,
        "rows": allrows,
        "dims": {
            "anio": sorted({r[0] for r in allrows}),
            "departamento": deptos,
            "categoria": ["Total", "Vacas", "Vaquillonas", "Novillos", "Novillitos",
                          "Terneros", "Terneras", "Toros", "Toritos", "Bueyes"],
        },
        "metricas": GAN_METRICAS,
        "notas": [
            "ganaderia: existencias bovinas al 31/12 de cada año (MAGyP), 2012–2025; "
            "'Salta' es el total provincial (suma de departamentos)."],
    }


# ==========================================================================
# TEMA — Minería: empleo minero (Salta) + recaudación tributaria (nacional)
# ==========================================================================
MINERIA_METRICAS = {
    "empleo_min":  {"label": "Empleo minero",              "unidad": "puestos",       "agg": "sum"},
    "recaud_ars":  {"label": "Recaudación (millones $)",   "unidad": "millones de $", "agg": "sum"},
    "recaud_usd":  {"label": "Recaudación (millones US$)", "unidad": "millones US$",  "agg": "sum"},
}
_MIN_IMP = {2: "Ganancias (sociedades)", 3: "Derechos de exportación", 4: "Seguridad social",
            5: "Créditos y débitos", 6: "Regalías y canon", 7: "Tasa ambiental"}


def _sq(v):
    """Quita comillas envolventes de los strings del bloque Empleo del xlsx."""
    return str(v).strip().strip('"').strip() if v is not None else v


def mineria():
    wb = openpyxl.load_workbook(_drv("CESS_Mineria.xlsx"), read_only=True, data_only=True)
    rows = []

    # ---- Empleo minero mensual (Salta) por rubro y género -------------------
    emp = list(wb["Empleo"].iter_rows(values_only=True))
    recs = []
    for r in emp[1:]:
        if not r or r[0] is None:
            continue
        fecha = r[0]
        anio, mes = fecha.year, fecha.month
        genero, rubro, cant = _sq(r[2]), _sq(r[3]), r[4]
        if cant is None:
            continue
        recs.append((anio, mes, _trim_label(anio, mes), genero, rubro, float(cant)))
    emp_df = pd.DataFrame(recs, columns=["anio", "mes", "trimestre", "genero", "rubro", "cant"])

    frames = []
    base = emp_df.groupby(["anio", "mes", "trimestre", "genero", "rubro"], as_index=False)["cant"].sum()
    frames.append(base.rename(columns={"genero": "g", "rubro": "r", "cant": "v"}))
    gt = emp_df.groupby(["anio", "mes", "trimestre", "rubro"], as_index=False)["cant"].sum()
    gt["g"] = "Total"; frames.append(gt.rename(columns={"rubro": "r", "cant": "v"}))
    rt = emp_df.groupby(["anio", "mes", "trimestre", "genero"], as_index=False)["cant"].sum()
    rt["r"] = "Total"; frames.append(rt.rename(columns={"genero": "g", "cant": "v"}))
    bt = emp_df.groupby(["anio", "mes", "trimestre"], as_index=False)["cant"].sum()
    bt["g"] = "Total"; bt["r"] = "Total"; frames.append(bt.rename(columns={"cant": "v"}))
    base = pd.concat(frames, ignore_index=True)

    def emit_emp(grano, keys, has_trim):
        g = base.groupby(keys, as_index=False)["v"].mean()
        for r in g.itertuples(index=False):
            d = r._asdict()
            rows.append([grano, int(d["anio"]), d["trimestre"] if has_trim else "",
                         d["g"], d["r"], "", "", "empleo_min", round(float(d["v"]), 1)])
    emit_emp("trimestral", ["g", "r", "anio", "trimestre"], True)
    emit_emp("anual", ["g", "r", "anio"], False)

    mc = emp_df.groupby("trimestre")["mes"].nunique()
    trims_completos = sorted([t for t, n in mc.items() if n >= 3], key=_trim_key)

    # ---- Recaudación tributaria anual por empresa e impuesto (nacional) -----
    imp = list(wb["Impuestos pagados 19-23"].iter_rows(values_only=True))
    for r in imp[1:]:
        if not r or r[0] is None:
            continue
        try:
            anio = int(r[0])
        except (TypeError, ValueError):
            continue
        empresa = str(r[1]).strip()
        for ci, label in _MIN_IMP.items():
            ars = r[ci]
            usd = r[ci + 12]   # bloque USD desplazado 12 columnas
            if ars is not None:
                rows.append(["anual", anio, "", "", "", empresa, label, "recaud_ars", round(float(ars), 3)])
            if usd is not None:
                rows.append(["anual", anio, "", "", "", empresa, label, "recaud_usd", round(float(usd), 3)])
    wb.close()

    fields = ["grano", "anio", "trimestre", "genero", "rubro", "empresa", "impuesto", "metrica", "valor"]
    anios = sorted({int(r[1]) for r in rows})
    trimestres = sorted({r[2] for r in rows if r[2]}, key=_trim_key)
    rubros = sorted({r[4] for r in rows if r[4] and r[4] != "Total"})
    empresas = sorted({r[5] for r in rows if r[5]})
    impuestos = [v for v in _MIN_IMP.values()]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": anios, "trimestre": trimestres,
            "genero": ["Total", "Femenino", "Masculino"],
            "rubro": rubros + ["Total"],
            "empresa": empresas, "impuesto": impuestos,
        },
        "metricas": MINERIA_METRICAS,
        "trims_completos": trims_completos,
        "notas": [
            "mineria: el empleo minero es de Salta (registro mensual desde 2007, media de meses por "
            "trimestre/año). La recaudación tributaria es NACIONAL del sector (por empresa, 2019–2023), "
            "no atribuible a Salta; el total en USD se recalcula como suma de impuestos.",
        ],
    }


# ==========================================================================
# TEMA — Financiero: préstamos y depósitos (BCRA) + inclusión financiera
# ==========================================================================
FIN_METRICAS = {
    "monto_corr": {"label": "Monto (pesos corrientes)", "unidad": "pesos", "agg": "sum"},
    "monto_real": {"label": "Monto (pesos constantes)", "unidad": "pesos", "agg": "sum"},
    "pda_10m":    {"label": "Puntos de acceso (cada 10.000 adultos)", "unidad": "PDA/10k", "agg": "mean"},
}
_FIN_BASE_Y = 2025


def financiero():
    ipc = _ipc_anual()
    base = ipc[_FIN_BASE_Y]

    def real(v, y):
        return v * base / ipc[y] if ipc.get(y) else None

    dp = pd.read_csv(_drv("BCRA_dep_pres_localidad_Salta.csv"), encoding="utf-8-sig")
    dp = dp[dp["moneda"] == "total"].copy()
    for c in ["anio", "trimestre", "prestamos_sector_privado", "depositos_sector_privado"]:
        dp[c] = pd.to_numeric(dp[c], errors="coerce")
    dp = dp.dropna(subset=["anio", "trimestre"])
    dp["anio"] = dp["anio"].astype(int)
    dp["trim_lbl"] = [f"{a}-T{int(t)}" for a, t in zip(dp["anio"], dp["trimestre"])]

    def geo(row):
        return "Salta" if row["nivel"] == "provincia" else norm_dept(row["departamento"])

    rows = []
    OPS = [("Préstamos", "prestamos_sector_privado"), ("Depósitos", "depositos_sector_privado")]
    sub = dp[dp["nivel"].isin(["departamento", "provincia"])].copy()
    sub["geo"] = sub.apply(geo, axis=1)
    for op, col in OPS:
        # Trimestral (nivel de stock a fin de trimestre)
        g = sub.groupby(["geo", "anio", "trim_lbl"], as_index=False)[col].sum()
        for r in g.itertuples(index=False):
            d = r._asdict()
            v = float(d[col]) * 1000.0; y = int(d["anio"])   # miles de $ -> $
            rows.append(["trimestral", y, d["trim_lbl"], d["geo"], op, "monto_corr", round(v, 1)])
            rr = real(v, y)
            if rr is not None:
                rows.append(["trimestral", y, d["trim_lbl"], d["geo"], op, "monto_real", round(rr, 1)])
        # Anual = promedio de los trimestres del año
        ga = sub.groupby(["geo", "anio", "trim_lbl"], as_index=False)[col].sum() \
                .groupby(["geo", "anio"], as_index=False)[col].mean()
        for r in ga.itertuples(index=False):
            d = r._asdict()
            v = float(d[col]) * 1000.0; y = int(d["anio"])   # miles de $ -> $
            rows.append(["anual", y, "", d["geo"], op, "monto_corr", round(v, 1)])
            rr = real(v, y)
            if rr is not None:
                rows.append(["anual", y, "", d["geo"], op, "monto_real", round(rr, 1)])

    # ---- Inclusión financiera: PDA por 10.000 adultos, por departamento -----
    inc = pd.read_csv(_drv("BCRA_IF_Salta_puntos_acceso_10000_adultos.csv"), encoding="utf-8-sig")
    inc["nu_pda_10m"] = pd.to_numeric(inc["nu_pda_10m"], errors="coerce")
    inc["anio"] = pd.to_numeric(inc["anio"], errors="coerce")
    inc = inc.dropna(subset=["nu_pda_10m", "anio"])
    inc["anio"] = inc["anio"].astype(int)
    inc["departamento"] = inc["tx_departamento"].map(norm_dept)
    # total de puntos = suma de tipos de PDA por (anio, mes, depto); anual = media de meses
    mensual = inc.groupby(["anio", "mes", "departamento"], as_index=False)["nu_pda_10m"].sum()
    anual = mensual.groupby(["anio", "departamento"], as_index=False)["nu_pda_10m"].mean()
    for r in anual.itertuples(index=False):
        rows.append(["anual", int(r.anio), "", r.departamento, "PDA", "pda_10m", round(float(r.nu_pda_10m), 2)])
    prov_inc = mensual.groupby(["anio", "mes"], as_index=False)["nu_pda_10m"].sum() \
                      .groupby("anio", as_index=False)["nu_pda_10m"].mean()
    for r in prov_inc.itertuples(index=False):
        rows.append(["anual", int(r.anio), "", "Salta", "PDA", "pda_10m", round(float(r.nu_pda_10m), 2)])

    fields = ["grano", "anio", "trimestre", "departamento", "operacion", "metrica", "valor"]
    deptos = [d for d in sorted({r[3] for r in rows}) if d not in NON_GEO and d not in PROVINCIAL_TOKENS]
    trimestres = sorted({r[2] for r in rows if r[2]}, key=_trim_key)
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({int(r[1]) for r in rows}),
            "trimestre": trimestres,
            "departamento": deptos,
            "operacion": ["Préstamos", "Depósitos", "PDA"],
        },
        "metricas": FIN_METRICAS,
        "trims_completos": trimestres,
        "notas": [
            "financiero: préstamos y depósitos al sector privado (BCRA, stock a fin de trimestre, "
            "miles de $); 'Salta' es el total provincial. Reales deflactados por IPC NOA (base 2025).",
            "financiero: inclusión financiera = puntos de acceso cada 10.000 adultos (promedio anual, "
            "suma de tipos); cobertura desde 2019.",
        ],
    }


# ==========================================================================
# TEMA — Permisos de construcción (INDEC), por municipio
# ==========================================================================
CONSTR_METRICAS = {
    "permisos":     {"label": "Permisos otorgados",   "unidad": "permisos", "agg": "sum"},
    "superficie_m2": {"label": "Superficie autorizada", "unidad": "m²",      "agg": "sum"},
    "share_sup_pct": {"label": "Participación en el país", "unidad": "%",     "agg": "mean"},
}


def _q_from_month_col(anio, mes_str):
    # mes_str tipo 'YYYY-MM-01'; devuelve (mes_lbl, trimestre)
    m = int(str(mes_str)[5:7])
    return f"{anio}-{m:02d}", _trim_label(anio, m)


def construccion():
    muni = pd.read_csv(_drv("Salta_permisos_construccion_INDEC_municipio_mensual.csv"), encoding="utf-8-sig")
    tot = pd.read_csv(_drv("Salta_permisos_construccion_INDEC_totales_mensual.csv"), encoding="utf-8-sig")
    for c in ["anio", "permisos_otorgados", "superficie_autorizada_m2"]:
        muni[c] = pd.to_numeric(muni[c], errors="coerce")
    muni = muni.dropna(subset=["anio"])
    muni["anio"] = muni["anio"].astype(int)
    # "Salta" (municipio) es la Ciudad de Salta: se renombra para no chocar con el token provincial.
    muni["municipio"] = muni["municipio"].map(lambda x: "Salta (ciudad)" if str(x).strip() == "Salta" else str(x).strip())
    muni["mes_lbl"] = [f"{a}-{int(str(m)[5:7]):02d}" for a, m in zip(muni["anio"], muni["mes"])]
    muni["trimestre"] = [_trim_label(a, int(str(m)[5:7])) for a, m in zip(muni["anio"], muni["mes"])]

    rows = []

    def emit_flows(df, geocol, geolabel_fn):
        MET = [("permisos", "permisos_otorgados"), ("superficie_m2", "superficie_autorizada_m2")]
        for met, col in MET:
            # mensual
            for r in df.itertuples(index=False):
                d = r._asdict()
                if pd.notna(d[col]):
                    rows.append(["mensual", int(d["anio"]), d["mes_lbl"], d["trimestre"],
                                 geolabel_fn(d), met, round(float(d[col]), 1)])
            # trimestral (suma) y anual (suma)
            gq = df.groupby([geocol, "anio", "trimestre"], as_index=False)[col].sum()
            for r in gq.itertuples(index=False):
                d = r._asdict()
                rows.append(["trimestral", int(d["anio"]), "", d["trimestre"], d[geocol], met, round(float(d[col]), 1)])
            ga = df.groupby([geocol, "anio"], as_index=False)[col].sum()
            for r in ga.itertuples(index=False):
                d = r._asdict()
                rows.append(["anual", int(d["anio"]), "", "", d[geocol], met, round(float(d[col]), 1)])

    emit_flows(muni, "municipio", lambda d: d["municipio"])

    # Provincial "Total Salta" = suma de municipios (por período). Se usa "Total Salta" porque
    # "Salta" ya es un municipio real (Ciudad de Salta).
    prov = muni.copy()
    prov["municipio"] = "Total Salta"
    emit_flows(prov, "municipio", lambda d: "Total Salta")

    # Participación nacional (del archivo de totales), provincial.
    tot["anio"] = pd.to_numeric(tot["anio"], errors="coerce").astype("Int64")
    tot["share_superficie_pct"] = pd.to_numeric(tot["share_superficie_pct"], errors="coerce")
    tot = tot.dropna(subset=["anio"])
    tot["mes_lbl"] = [f"{int(a)}-{int(str(m)[5:7]):02d}" for a, m in zip(tot["anio"], tot["mes"])]
    tot["trimestre"] = [_trim_label(int(a), int(str(m)[5:7])) for a, m in zip(tot["anio"], tot["mes"])]
    for r in tot.itertuples(index=False):
        d = r._asdict()
        if pd.notna(d["share_superficie_pct"]):
            rows.append(["mensual", int(d["anio"]), d["mes_lbl"], d["trimestre"], "Total Salta", "share_sup_pct", round(float(d["share_superficie_pct"]), 2)])
    tq = tot.groupby(["anio", "trimestre"], as_index=False)["share_superficie_pct"].mean()
    for r in tq.itertuples(index=False):
        rows.append(["trimestral", int(r.anio), "", r.trimestre, "Total Salta", "share_sup_pct", round(float(r.share_superficie_pct), 2)])
    ty = tot.groupby("anio", as_index=False)["share_superficie_pct"].mean()
    for r in ty.itertuples(index=False):
        rows.append(["anual", int(r.anio), "", "", "Total Salta", "share_sup_pct", round(float(r.share_superficie_pct), 2)])

    # No emitir trimestres/años incompletos (último período parcial => caída ficticia).
    trims_ok, anios_ok, trims_completos = _completos(muni, mes_col="mes_lbl")
    rows = _keep_completos(rows, trims_ok, anios_ok, gi=0, ai=1, ti=3)

    fields = ["grano", "anio", "mes", "trimestre", "municipio", "metrica", "valor"]
    munis = [m for m in sorted({r[4] for r in rows}) if m not in NON_GEO and m not in PROVINCIAL_TOKENS]
    trimestres = sorted({r[3] for r in rows if r[3]}, key=_trim_key)
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({int(r[1]) for r in rows}),
            "mes": sorted({r[2] for r in rows if r[2]}),
            "trimestre": trimestres,
            "municipio": munis,
        },
        "metricas": CONSTR_METRICAS,
        "trims_completos": trims_completos,
        "notas": [
            "construccion: permisos de edificación privada informados a INDEC por 5 municipios de "
            "Salta (mensual); los flujos se SUMAN por trimestre/año. La superficie es intención de "
            "construir, no obra ejecutada; 'Salta' suma los municipios."],
    }


# ==========================================================================
# TEMA — Recursos a municipios (Contaduría Gral. de Salta)
# ==========================================================================
RECURSOS_METRICAS = {
    "monto_corr": {"label": "Monto (pesos corrientes)", "unidad": "pesos", "agg": "sum"},
    "monto_real": {"label": "Monto (pesos constantes)", "unidad": "pesos", "agg": "sum"},
}
_REC_BASE_Y = 2025


def recursos_municipios():
    ipc = _ipc_anual()
    base = ipc[_REC_BASE_Y]

    def real(v, y):
        return v * base / ipc[y] if ipc.get(y) else None

    df = pd.read_csv(_drv("Salta_recursos_municipios_mensual.csv"), encoding="utf-8-sig")
    df["departamento"] = df["departamento"].map(norm_dept)
    # El municipio "Salta" es la Ciudad de Salta: se renombra para no chocar con el token provincial.
    df["municipio"] = df["municipio"].map(lambda x: "Salta (ciudad)" if str(x).strip() == "Salta" else str(x).strip())
    df["anio"] = pd.to_numeric(df["anio"], errors="coerce").astype("Int64")
    df["mes"] = pd.to_numeric(df["mes"], errors="coerce").astype("Int64")
    df["monto"] = pd.to_numeric(df["monto"], errors="coerce")
    df = df.dropna(subset=["anio", "mes", "monto"])
    df["anio"] = df["anio"].astype(int)
    df["trimestre"] = [_trim_label(a, m) for a, m in zip(df["anio"], df["mes"])]

    rows = []

    def emit(keycols, group):
        # trimestral y anual (suma de meses); corrientes + reales
        gq = group.groupby(keycols + ["anio", "trimestre"], as_index=False)["monto"].sum()
        for r in gq.itertuples(index=False):
            d = r._asdict(); y = int(d["anio"]); v = float(d["monto"])
            rows.append(["trimestral", y, d["trimestre"], d["departamento"], d["municipio"], d["grupo"], "monto_corr", round(v, 1)])
            rr = real(v, y)
            if rr is not None:
                rows.append(["trimestral", y, d["trimestre"], d["departamento"], d["municipio"], d["grupo"], "monto_real", round(rr, 1)])
        ga = group.groupby(keycols + ["anio"], as_index=False)["monto"].sum()
        for r in ga.itertuples(index=False):
            d = r._asdict(); y = int(d["anio"]); v = float(d["monto"])
            rows.append(["anual", y, "", d["departamento"], d["municipio"], d["grupo"], "monto_corr", round(v, 1)])
            rr = real(v, y)
            if rr is not None:
                rows.append(["anual", y, "", d["departamento"], d["municipio"], d["grupo"], "monto_real", round(rr, 1)])

    emit(["departamento", "municipio", "grupo"], df)

    # No emitir períodos incompletos (último trimestre/año parcial => caída ficticia).
    trims_ok, anios_ok, trims_ok_list = _completos(df, mes_col="mes")
    rows = _keep_completos(rows, trims_ok, anios_ok)

    fields = ["grano", "anio", "trimestre", "departamento", "municipio", "grupo", "metrica", "valor"]
    deptos = [d for d in sorted({r[3] for r in rows}) if d not in NON_GEO and d not in PROVINCIAL_TOKENS]
    munis = [m for m in sorted({r[4] for r in rows}) if m and m not in PROVINCIAL_TOKENS]
    trimestres = sorted({r[2] for r in rows if r[2]}, key=_trim_key)
    grupos = sorted({r[5] for r in rows if r[5]})
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({int(r[1]) for r in rows}),
            "trimestre": trimestres,
            "departamento": deptos,
            "municipio": munis,
            "grupo": grupos,
        },
        "metricas": RECURSOS_METRICAS,
        "trims_completos": trims_ok_list,
        "notas": [
            "recursos-municipios: transferencias a municipios de Salta (Contaduría Gral.), mensual "
            "desde 2021, sumadas por trimestre/año. Reales deflactados por IPC NOA (base 2025). "
            "Grupos: Coparticipación, Regalías, Canon, Fondo compensador, Otros."],
    }


# ==========================================================================
# TEMA — Energía eléctrica renovable (CAMMESA): generación por fuente y potencia
# ==========================================================================
ENERGIA_METRICAS = {
    "generacion_gwh":      {"label": "Generación",           "unidad": "GWh", "agg": "sum"},
    "gen_central_gwh":     {"label": "Generación por central", "unidad": "GWh", "agg": "sum"},
    "share_renovable_pct": {"label": "Participación renovable", "unidad": "%",  "agg": "mean"},
    "potencia_mw":         {"label": "Potencia instalada",    "unidad": "MW",  "agg": "sum"},
}
_ENE_FUENTE = {"hidraulica": "Hidráulica", "renovable": "Renovable", "termica": "Térmica"}


def energia_renovable():
    rows = []

    # ---- Generación por fuente (archivo ancho, MWh -> GWh) ------------------
    fu = pd.read_csv(_drv("Salta_generacion_CAMMESA_fuente_mensual.csv"), encoding="utf-8-sig")
    fu["anio"] = pd.to_numeric(fu["anio"], errors="coerce").astype("Int64")
    fu = fu.dropna(subset=["anio"])
    fu["anio"] = fu["anio"].astype(int)
    fu["m"] = fu["mes"].astype(str).str.slice(5, 7).astype(int)
    fu["trimestre"] = [_trim_label(a, m) for a, m in zip(fu["anio"], fu["m"])]
    long = fu.melt(id_vars=["anio", "trimestre"], value_vars=list(_ENE_FUENTE.keys()),
                   var_name="fu", value_name="v")
    long["v"] = pd.to_numeric(long["v"], errors="coerce") / 1000.0   # MWh -> GWh
    long["fuente"] = long["fu"].map(_ENE_FUENTE)
    long = long.dropna(subset=["v"])
    gq = long.groupby(["fuente", "anio", "trimestre"], as_index=False)["v"].sum()
    for r in gq.itertuples(index=False):
        rows.append(["trimestral", int(r.anio), r.trimestre, r.fuente, "", "generacion_gwh", round(float(r.v), 1)])
    ga = long.groupby(["fuente", "anio"], as_index=False)["v"].sum()
    for r in ga.itertuples(index=False):
        rows.append(["anual", int(r.anio), "", r.fuente, "", "generacion_gwh", round(float(r.v), 1)])

    # ---- % renovable (media del período) -----------------------------------
    fu["share_renovable_pct"] = pd.to_numeric(fu["share_renovable_pct"], errors="coerce")
    sq = fu.groupby(["anio", "trimestre"], as_index=False)["share_renovable_pct"].mean()
    for r in sq.itertuples(index=False):
        rows.append(["trimestral", int(r.anio), r.trimestre, "", "", "share_renovable_pct", round(float(r.share_renovable_pct), 1)])
    sa = fu.groupby("anio", as_index=False)["share_renovable_pct"].mean()
    for r in sa.itertuples(index=False):
        rows.append(["anual", int(r.anio), "", "", "", "share_renovable_pct", round(float(r.share_renovable_pct), 1)])

    # Completitud a partir de la cobertura mensual por fuente.
    _trims_ok, _anios_ok, trims_completos = _completos(fu, mes_col="m")

    # ---- Generación por central (MWh -> GWh), anual ------------------------
    ce = pd.read_csv(_drv("Salta_generacion_CAMMESA_central_mensual.csv"), encoding="utf-8-sig")
    ce["anio"] = pd.to_numeric(ce["anio"], errors="coerce").astype("Int64")
    ce = ce.dropna(subset=["anio"])
    ce["anio"] = ce["anio"].astype(int)
    ce["gwh"] = pd.to_numeric(ce["generacion_neta_mwh"], errors="coerce") / 1000.0
    ce["central"] = ce["agente_descripcion"].astype(str).str.strip()
    gc = ce.groupby(["central", "anio"], as_index=False)["gwh"].sum()
    for r in gc.itertuples(index=False):
        rows.append(["anual", int(r.anio), "", "", r.central, "gen_central_gwh", round(float(r.gwh), 1)])

    # ---- Potencia instalada por central (snapshot, MW) ---------------------
    po = pd.read_csv(_drv("Salta_generacion_CAMMESA_potencia_instalada.csv"), encoding="utf-8-sig")
    po["mw"] = pd.to_numeric(po["potencia_instalada_mw"], errors="coerce")
    po["central"] = po["agente_descripcion"].astype(str).str.strip()
    po["anio"] = pd.to_numeric(po["anio"], errors="coerce").astype("Int64")
    po = po.dropna(subset=["mw"])
    pot_anio = int(po["anio"].max())
    pg = po.groupby("central", as_index=False)["mw"].sum()
    for r in pg.itertuples(index=False):
        rows.append(["anual", pot_anio, "", "", r.central, "potencia_mw", round(float(r.mw), 1)])

    # No emitir períodos incompletos: trimestres parciales (todas las series) y años parciales
    # SOLO para los flujos de generación (la potencia es una foto del último año y se conserva).
    _GEN_ANUAL = {"generacion_gwh", "gen_central_gwh"}
    rows = [r for r in rows
            if not (r[0] == "trimestral" and r[2] not in _trims_ok)
            and not (r[0] == "anual" and r[5] in _GEN_ANUAL and int(r[1]) not in _anios_ok)]

    fields = ["grano", "anio", "trimestre", "fuente", "central", "metrica", "valor"]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({int(r[1]) for r in rows}),
            "trimestre": trims_completos,
            "fuente": ["Hidráulica", "Renovable", "Térmica"],
            "central": sorted({r[4] for r in rows if r[4]}),
        },
        "metricas": ENERGIA_METRICAS,
        "trims_completos": trims_completos,
        "notas": [
            "energia-renovable: generación eléctrica de Salta (CAMMESA), MWh convertidos a GWh; "
            "los flujos se suman por trimestre/año. 'Renovable' = régimen Ley 27.191; la potencia "
            "instalada es una foto al último mes disponible."],
    }


# ==========================================================================
# TEMA — Inversión en I+D e innovación (RICyT/MINCyT), provincial Salta
#   Dataset chico y estático (Sheet sin script ETL); se embebe verificado.
# ==========================================================================
ID_METRICAS = {
    "inv_id_corr": {"label": "Inversión en I+D (pesos corrientes)", "unidad": "millones de $", "agg": "sum"},
    "inv_id_real": {"label": "Inversión en I+D (pesos de 2017)",   "unidad": "millones de $", "agg": "sum"},
    "personal_id": {"label": "Personal en I+D",                     "unidad": "personas",      "agg": "sum"},
}
# Inversión I+D de Salta: anio -> (millones $ corrientes, millones $ constantes 2017)
_ID_INV = {2017: (678, 678), 2018: (859, 640), 2019: (1083, 528), 2020: (1199, 416),
           2021: (1909, 451), 2022: (3400, 464), 2023: (8196, 478), 2024: (22608, 408)}
# Personal en I+D por función (headcount): anio -> (investigadores/as y becarios/as, técnicos y apoyo)
_ID_PERS = {
    2003: (462, None), 2004: (489, None), 2005: (508, None), 2006: (679, None),
    2007: (708, None), 2008: (730, None), 2009: (710, 296), 2010: (648, 282),
    2011: (691, 278), 2012: (674, 324), 2013: (672, 344), 2014: (762, 349),
    2015: (914, 322), 2016: (865, 333), 2017: (868, 323), 2018: (920, 293),
    2019: (923, 279), 2020: (887, 298), 2021: (911, 315), 2022: (919, 342),
    2023: (934, 373), 2024: (925, 382),
}


def id_innovacion():
    rows = []
    for a, (corr, real) in _ID_INV.items():
        rows.append([a, "", "inv_id_corr", float(corr)])
        rows.append([a, "", "inv_id_real", float(real)])
    for a, (inv, tec) in _ID_PERS.items():
        if inv is not None:
            rows.append([a, "Investigadores y becarios", "personal_id", float(inv)])
        if tec is not None:
            rows.append([a, "Personal técnico y de apoyo", "personal_id", float(tec)])

    fields = ["anio", "funcion", "metrica", "valor"]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({r[0] for r in rows}),
            "funcion": ["Investigadores y becarios", "Personal técnico y de apoyo"],
        },
        "metricas": ID_METRICAS,
        "notas": [
            "id-innovacion: inversión provincial en I+D (millones de $ corrientes y en pesos "
            "constantes de 2017) y personal en I+D por función; fuente RICyT/MINCyT. En términos "
            "reales la inversión se mantuvo casi estancada (678 en 2017 → 408 en 2024)."],
    }


# ==========================================================================
# TEMA — Resultado fiscal provincial (Esquema Ahorro-Inversión-Financiamiento)
# ==========================================================================
RESFISC_METRICAS = {
    "monto_corr": {"label": "Monto (pesos corrientes)", "unidad": "pesos", "agg": "sum"},
    "monto_real": {"label": "Monto (pesos constantes)", "unidad": "pesos", "agg": "sum"},
    # Cociente contra el gasto primario. Es un RATIO: no depende de la deflación (verificado
    # período por período, corrientes y constantes dan el mismo valor), así que no se duplica
    # en corr/real. agg "mean" es defensivo: sumar los porcentajes de dos conceptos no
    # significaría nada, y los gráficos siempre fijan o abren el concepto en series.
    "pct_gprim": {"label": "En % del gasto primario", "unidad": "% del gasto primario",
                  "agg": "mean"},
}
_RESFISC_BASE_Y = 2025
# (etiqueta legible, prefijo de columna en el CSV). XI == VIII en Salta: se usa XI como
# "Resultado financiero". El primario es cálculo propio (XI + intereses de deuda).
_RESFISC_CONCEPTOS = [
    ("Ingresos totales",      "ingresos_totales"),
    ("Gastos totales",        "gastos_totales"),
    ("Resultado financiero",  "res_financiero_xi"),
    ("Resultado primario",    "res_primario"),
    ("Intereses de la deuda", "intereses_deuda"),
]
# Gasto primario = gastos totales − intereses de la deuda. La identidad se verificó contra la
# propia fuente: ingresos − gasto primario == resultado primario, exacto en todos los períodos.
_RESFISC_GPRIM = "Gasto primario"
# Conceptos que se expresan como % del gasto primario.
_RESFISC_PCT = ["Resultado financiero", "Resultado primario"]


def resultado_fiscal():
    ipcm = _ipc_noa_map()
    base = _ipc_anual()[_RESFISC_BASE_Y]
    # thousands="," tolera exportaciones con separador de miles entrecomillado
    # ("12,021,478,627"); en el CSV limpio (decimal con punto) es inocuo.
    df = pd.read_csv(_drv("Resultado_Fiscal_Salta_AIF_mensual_desde_2021.csv"),
                     encoding="utf-8-sig", thousands=",")
    for c in df.columns:
        if c not in ("provincia", "periodo", "fecha_corte"):
            df[c] = pd.to_numeric(df[c], errors="coerce")

    rows = []
    for _, r in df.iterrows():
        if pd.isna(r["anio"]) or pd.isna(r["mes"]):
            continue
        y, m = int(r["anio"]), int(r["mes"])
        periodo = f"{y}-{m:02d}"
        ipc = ipcm.get(y * 100 + m)
        for vista, suf in (("acumulado", "acum"), ("mensual", "mensual")):
            vals = {}
            for concepto, pref in _RESFISC_CONCEPTOS:
                v = r.get(f"{pref}_{suf}")
                if v is None or pd.isna(v):
                    continue
                v = float(v)
                vals[concepto] = v
                rows.append([vista, y, m, periodo, concepto, "monto_corr", round(v, 1)])
                if ipc:
                    rows.append([vista, y, m, periodo, concepto, "monto_real", round(v * base / ipc, 1)])

            # Gasto primario y los resultados expresados en % de ese gasto.
            gt, it = vals.get("Gastos totales"), vals.get("Intereses de la deuda")
            if gt is None or it is None:
                continue
            gprim = gt - it
            rows.append([vista, y, m, periodo, _RESFISC_GPRIM, "monto_corr", round(gprim, 1)])
            if ipc:
                rows.append([vista, y, m, periodo, _RESFISC_GPRIM, "monto_real",
                             round(gprim * base / ipc, 1)])
            if gprim:
                for concepto in _RESFISC_PCT:
                    if concepto in vals:
                        rows.append([vista, y, m, periodo, concepto, "pct_gprim",
                                     round(vals[concepto] / gprim * 100, 2)])

    fields = ["grano", "anio", "mes", "periodo", "concepto", "metrica", "valor"]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "grano": ["acumulado", "mensual"],
            "anio": sorted({r[1] for r in rows}),
            "periodo": sorted({r[3] for r in rows}),
            "concepto": [c for c, _ in _RESFISC_CONCEPTOS] + [_RESFISC_GPRIM],
        },
        "metricas": RESFISC_METRICAS,
        "notas": [
            "resultado-fiscal: Esquema Ahorro-Inversión-Financiamiento (ejecución consolidada Adm. "
            "Central + Organismos Descentralizados, devengado), fuente presupuesto.salta.gob.ar. "
            "Resultado financiero = ingresos totales − gastos totales (VIII = XI en Salta).",
            "resultado-fiscal: el resultado primario es un cálculo propio (resultado financiero + "
            "intereses de la deuda), la fuente no publica una línea primaria. Reales deflactados por "
            "IPC NOA (base 2025).",
            "resultado-fiscal: `pct_gprim` expresa el resultado financiero y el primario como % del "
            "GASTO PRIMARIO (gastos totales − intereses de la deuda). Es la medida de esfuerzo fiscal "
            "que no depende de la inflación ni del tamaño nominal del presupuesto. Al ser un cociente "
            "entre dos flujos del mismo período da idéntico en pesos corrientes y constantes, así que "
            "no se duplica por moneda. En el acumulado los meses NO son comparables entre sí (enero "
            "arranca alto y el ratio baja al avanzar el año): la comparación válida es contra el mismo "
            "mes del año anterior.",
            "resultado-fiscal: el informe de diciembre es el cierre anual y reexpresa el resultado del "
            "año (el flujo mensual de diciembre incorpora ajustes de cierre); 2024-12 no fue publicado.",
        ],
    }


# ==========================================================================
# TEMA — Recaudación de Ingresos Brutos (Convenio Multilateral), por sector
# ==========================================================================
RECAUD_METRICAS = {
    "recaud_corr": {"label": "Recaudación (pesos corrientes)", "unidad": "pesos", "agg": "sum"},
    "recaud_real": {"label": "Recaudación (pesos constantes)", "unidad": "pesos", "agg": "sum"},
}
_RECAUD_BASE_Y = 2025
# Etiqueta legible por letra CIIU (sección de actividad).
_SECTOR_LABEL = {
    "A": "Agro, ganadería y silvicultura", "B": "Pesca", "C": "Minas y canteras",
    "D": "Industria manufacturera", "E": "Electricidad, gas y agua", "F": "Construcción",
    "G": "Comercio", "H": "Hotelería y restaurantes", "I": "Transporte y comunicaciones",
    "J": "Intermediación financiera", "K": "Servicios inmobiliarios y empresariales",
    "L": "Administración pública", "M": "Enseñanza", "N": "Salud y servicios sociales",
    "O": "Servicios comunitarios y personales", "P": "Servicio doméstico",
    "Q": "Organizaciones extraterritoriales",
}


def _sector_label(letra):
    key = str(letra).strip().upper()
    if key.startswith("SIN"):
        return "Sin clasificar"
    return _SECTOR_LABEL.get(key, str(letra).strip().title())


def recaudacion():
    ipcm = _ipc_noa_map()
    base = _ipc_anual()[_RECAUD_BASE_Y]
    df = pd.read_csv(_drv("Recaudacion_AAEE_ConvenioMultilateral_Salta_por_letra_desde_2021.csv"),
                     encoding="utf-8-sig", thousands=",")
    df["anio"] = pd.to_numeric(df["anio"], errors="coerce").astype("Int64")
    df["mes"] = pd.to_numeric(df["mes"], errors="coerce").astype("Int64")
    df["recaudacion_ars"] = pd.to_numeric(df["recaudacion_ars"], errors="coerce")
    df = df.dropna(subset=["anio", "mes", "recaudacion_ars"])
    df["anio"] = df["anio"].astype(int)
    df["mes"] = df["mes"].astype(int)
    df["sector"] = df["letra"].map(_sector_label)
    df["trimestre"] = [_trim_label(a, m) for a, m in zip(df["anio"], df["mes"])]
    df["ipc"] = (df["anio"] * 100 + df["mes"]).map(ipcm)
    df["real"] = df["recaudacion_ars"] * base / df["ipc"]

    rows = []
    # Mensual (cada mes, por sector)
    for r in df.itertuples(index=False):
        y, m = int(r.anio), int(r.mes)
        periodo = f"{y}-{m:02d}"
        rows.append(["mensual", y, m, r.trimestre, periodo, r.sector, "recaud_corr", round(float(r.recaudacion_ars), 1)])
        if pd.notna(r.real):
            rows.append(["mensual", y, m, r.trimestre, periodo, r.sector, "recaud_real", round(float(r.real), 1)])
    # Trimestral y anual (suma de meses), por sector
    gq = df.groupby(["sector", "anio", "trimestre"], as_index=False).agg(corr=("recaudacion_ars", "sum"), real=("real", "sum"))
    for r in gq.itertuples(index=False):
        rows.append(["trimestral", int(r.anio), 0, r.trimestre, "", r.sector, "recaud_corr", round(float(r.corr), 1)])
        rows.append(["trimestral", int(r.anio), 0, r.trimestre, "", r.sector, "recaud_real", round(float(r.real), 1)])
    ga = df.groupby(["sector", "anio"], as_index=False).agg(corr=("recaudacion_ars", "sum"), real=("real", "sum"))
    for r in ga.itertuples(index=False):
        rows.append(["anual", int(r.anio), 0, "", "", r.sector, "recaud_corr", round(float(r.corr), 1)])
        rows.append(["anual", int(r.anio), 0, "", "", r.sector, "recaud_real", round(float(r.real), 1)])

    # No emitir el último trimestre/año incompleto (caída ficticia).
    trims_ok, anios_ok, _tl = _completos(df, mes_col="mes")
    rows = _keep_completos(rows, trims_ok, anios_ok, gi=0, ai=1, ti=3)

    fields = ["grano", "anio", "mes", "trimestre", "periodo", "sector", "metrica", "valor"]
    sectores = [s for s in _SECTOR_LABEL.values()] + ["Sin clasificar"]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "grano": ["mensual", "trimestral", "anual"],
            "anio": sorted({r[1] for r in rows if r[0] == "anual"}),
            "trimestre": sorted({r[3] for r in rows if r[0] == "trimestral" and r[3]}, key=_trim_key),
            "periodo": sorted({r[4] for r in rows if r[4]}),
            "sector": [s for s in sectores if s in {r[5] for r in rows}],
        },
        "metricas": RECAUD_METRICAS,
        "notas": [
            "recaudacion: Impuesto a las Actividades Económicas (Ingresos Brutos), régimen CONVENIO "
            "MULTILATERAL únicamente (no incluye contribuyentes locales/directos); por sector de "
            "actividad (CIIU). Fuente: DGR / Ministerio de Economía de Salta.",
            "recaudacion: mensual desde 2021, sumada por trimestre/año; reales deflactados por IPC "
            "NOA (base 2025). No se emite el último período incompleto.",
        ],
    }


# ==========================================================================
# TEMA — Exportaciones de Salta por rubro y país de destino (INDEC COMEX)
# ==========================================================================
# Valor FOB en dólares CORRIENTES: no se deflacta (no hay deflactor de comercio exterior en
# el repo y el estándar del INDEC es dólares corrientes). Se emite en USD enteros y en
# toneladas, NO en millones: el formateador compacto de charts.js abrevia 1447,687 como
# "1,4 K" (ambiguo, son 1.447 millones) mientras que 1.447.687.147 lo abrevia "1447,7 M".
EXPO_METRICAS = {
    "fob_usd": {"label": "Exportaciones (USD FOB)", "unidad": "USD", "agg": "sum"},
    "peso_ton": {"label": "Volumen exportado (toneladas)", "unidad": "toneladas", "agg": "sum"},
}

# Primer dígito del código de país del INDEC -> continente. Verificado contra los 146 países
# presentes (cada bloque trae además su propio "Indeterminado (América)", "(Europa)", etc.).
_EXPO_CONTINENTE = {
    "1": "África", "2": "América", "3": "Asia", "4": "Europa", "5": "Oceanía",
    "9": "Indeterminado",
}
_EXPO_CONT_ORDEN = ["América", "Asia", "Europa", "África", "Oceanía", "Indeterminado"]

# Etiquetas más cortas para las leyendas: las del CSV son las del INDEC, demasiado largas.
_EXPO_GR_LABEL = {
    "Manufacturas de origen agropecuario (MOA)": "Manufacturas agropecuarias (MOA)",
    "Manufacturas de origen industrial (MOI)": "Manufacturas industriales (MOI)",
}
_EXPO_GR_ORDEN = ["Productos primarios", "Manufacturas agropecuarias (MOA)",
                  "Manufacturas industriales (MOI)", "Combustibles y energía"]
# Los rubros x9899 son el agregado por secreto estadístico: el INDEC no publica el rubro fino,
# sólo el gran rubro. Se muestran como categoría propia (no se reparten ni se excluyen), así
# que necesitan una etiqueta corta para el ranking.
_EXPO_RUBRO_LABEL = {
    "Productos primarios (confidencial)": "Confidencial (primarios)",
    "Manufacturas de origen agropecuario (confidencial)": "Confidencial (MOA)",
    "Manufacturas de origen industrial (confidencial)": "Confidencial (MOI)",
    "Combustibles y energía (confidencial)": "Confidencial (combustibles)",
    "Resto de los productos de molinería y de las preparaciones a base de cereales,harina,"
    "almidón,fécula o leche,productos de pastelería": "Resto de molinería y panificados",
    "Resto de residuos alimenticios y preparados para animales": "Resto de alimento animal",
    "Resto de hortalizas y legumbres sin elaborar": "Resto de hortalizas sin elaborar",
    "Resto de azúcar y artículos de confitería": "Resto de azúcar y confitería",
    "Resto semillas y frutos oleaginosos": "Resto de semillas oleaginosas",
}


def exportaciones():
    df = pd.read_csv(_drv("Salta_exportaciones_INDEC_rubro_destino_desde_2021.csv"),
                     encoding="utf-8-sig")
    df["anio"] = pd.to_numeric(df["anio"], errors="coerce").astype("Int64")
    df["fob_usd"] = pd.to_numeric(df["fob_usd"], errors="coerce")
    df["peso_neto_kg"] = pd.to_numeric(df["peso_neto_kg"], errors="coerce")
    df = df.dropna(subset=["anio", "fob_usd"])
    df["anio"] = df["anio"].astype(int)

    df["gran_rubro"] = [_EXPO_GR_LABEL.get(g, g) for g in df["gran_rubro"].astype(str).str.strip()]
    ru = df["rubro"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()
    df["rubro"] = [_EXPO_RUBRO_LABEL.get(r, r) for r in ru]
    df["continente"] = df["pais_cod"].astype(str).str.strip().str[0].map(_EXPO_CONTINENTE)
    faltan_cont = int(df["continente"].isna().sum())
    df["continente"] = df["continente"].fillna("Indeterminado")

    # Sólo filas atómicas (anio x rubro x país): ningún total precalculado, así el navegador
    # agrega por gran rubro / continente / país / rubro sin riesgo de doble conteo.
    rows = []
    for r in df.itertuples(index=False):
        rows.append([r.anio, r.gran_rubro, r.rubro, r.continente, r.pais_destino,
                     "fob_usd", round(float(r.fob_usd), 2)])
        if pd.notna(r.peso_neto_kg):
            rows.append([r.anio, r.gran_rubro, r.rubro, r.continente, r.pais_destino,
                         "peso_ton", round(float(r.peso_neto_kg) / 1000.0, 3)])

    fields = ["anio", "gran_rubro", "rubro", "continente", "pais_destino", "metrica", "valor"]
    tot = df.groupby("anio")["fob_usd"].sum()
    conf = df[df["confidencial"] == 1].groupby("anio")["fob_usd"].sum()
    pct = (conf / tot * 100).round(1).dropna()
    notas = [
        "exportaciones: exportaciones de Salta por rubro y país de destino, base ANUAL del INDEC "
        "(COMEX, `Datos_origen_2002_2025.xlsb`, provincia de ORIGEN). Valor FOB en dólares "
        "CORRIENTES (sin deflactar) y peso neto. Metodología y controles en "
        "`datos-drive/Metodologia_Salta_exportaciones_INDEC_rubro_destino.md`.",
        "exportaciones: el secreto estadístico oculta el rubro fino de una parte creciente del "
        "valor exportado (" + "; ".join("%d: %s %%" % (a, p) for a, p in pct.items()) + "). Esas "
        "operaciones se muestran como categoría propia ('Confidencial (…)'), conservan país de "
        "destino y gran rubro, y NO se reparten entre los rubros identificados.",
        "exportaciones: `gran_rubro` y `continente` se derivan del primer dígito del código de "
        "rubro y de país del INDEC. El máximo detalle de producto con apertura provincial es el "
        "rubro; la posición arancelaria (NCM) sólo existe a nivel nacional.",
        "exportaciones: el control automático de saltos de magnitud marca valores a >1000× de la "
        "mediana. NO es un artefacto de parseo: la fuente es binaria (.xlsb leído con pyxlsb, sin "
        "separadores que interpretar) y la distribución del comercio es de cola muy larga (la "
        "mediana de una celda rubro×país ronda los 135.000 USD y la mayor supera los 380 M). "
        "Verificado además contra el total nacional del INDEC, con diferencia de 0,98 USD sobre "
        "87.111 M.",
    ]
    if faltan_cont:
        notas.append("exportaciones: %d filas con código de país sin continente asignable; "
                     "se agrupan en 'Indeterminado'." % faltan_cont)
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({r[0] for r in rows}),
            "gran_rubro": [g for g in _EXPO_GR_ORDEN if g in {r[1] for r in rows}],
            "rubro": sorted({r[2] for r in rows}),
            "continente": [c for c in _EXPO_CONT_ORDEN if c in {r[3] for r in rows}],
            "pais_destino": sorted({r[4] for r in rows}),
        },
        "metricas": EXPO_METRICAS,
        "notas": notas,
    }


# ==========================================================================
# TEMA — Salud (estadísticas vitales y de servicios, MSP Salta)
# ==========================================================================
# Cuatro CSV con unidades de análisis distintas se funden en un solo esquema canónico:
#   provincial            -> departamento="Salta", area_operativa="Total"
#   por Área Operativa    -> departamento=<depto>, area_operativa=<AO>
#   por departamento      -> departamento=<depto>, area_operativa="Total"
# `nacidos_vivos` (provincial, con apertura por peso) y `nacidos_vivos_res` (por residencia)
# son métricas DISTINTAS a propósito: comparten el concepto pero no la unidad de análisis, y
# mezclarlas haría que un gráfico sin filtro sumara la serie provincial con sus propias aperturas.
SALUD_METRICAS = {
    "nacidos_vivos":         {"label": "Nacidos vivos",                 "unidad": "nacimientos", "agg": "sum"},
    "nacidos_vivos_res":     {"label": "Nacidos vivos (por residencia)", "unidad": "nacimientos", "agg": "sum"},
    "defunciones":           {"label": "Defunciones",                   "unidad": "defunciones", "agg": "sum"},
    "defunciones_infantiles": {"label": "Defunciones de menores de 1 año", "unidad": "defunciones", "agg": "sum"},
    "defunciones_maternas":  {"label": "Defunciones maternas",          "unidad": "defunciones", "agg": "sum"},
    "defunciones_fetales":   {"label": "Defunciones fetales",           "unidad": "defunciones", "agg": "sum"},
    "otras_defunciones":     {"label": "Otras defunciones",             "unidad": "defunciones", "agg": "sum"},
    "matrimonios":           {"label": "Matrimonios",                   "unidad": "matrimonios", "agg": "sum"},
    "tasa_natalidad":        {"label": "Tasa de natalidad",             "unidad": "por mil habitantes",    "agg": "mean"},
    "tasa_mortalidad_general": {"label": "Tasa de mortalidad general",  "unidad": "por mil habitantes",    "agg": "mean"},
    "tasa_mortalidad_infantil": {"label": "Tasa de mortalidad infantil", "unidad": "por mil nacidos vivos", "agg": "mean"},
    "tasa_mortalidad_materna": {"label": "Tasa de mortalidad materna",  "unidad": "por diez mil nacidos vivos", "agg": "mean"},
    "poblacion":             {"label": "Población estimada",            "unidad": "habitantes", "agg": "mean"},
    "consultas_medicas":     {"label": "Consultas médicas",             "unidad": "consultas",  "agg": "sum"},
    "consultas_por_habitante": {"label": "Consultas médicas por habitante", "unidad": "consultas/hab.", "agg": "mean"},
    "camas_disponibles":     {"label": "Camas disponibles",             "unidad": "camas (promedio)", "agg": "sum"},
    "pacientes_dia":         {"label": "Pacientes-día",                 "unidad": "pacientes (promedio)", "agg": "sum"},
    "ocupacion_camas":       {"label": "Ocupación de camas",            "unidad": "%",          "agg": "mean"},
    "permanencia_promedio":  {"label": "Permanencia promedio",          "unidad": "días",       "agg": "mean"},
    "giro_camas":            {"label": "Giro de camas",                 "unidad": "egresos/cama", "agg": "mean"},
    "egresos":               {"label": "Egresos hospitalarios",         "unidad": "egresos",    "agg": "sum"},
    "altas":                 {"label": "Altas",                         "unidad": "altas",      "agg": "sum"},
    "defunciones_hosp":      {"label": "Defunciones hospitalarias",     "unidad": "defunciones", "agg": "sum"},
    "pases":                 {"label": "Pases internos",                "unidad": "pases",      "agg": "sum"},
    "establecimientos":      {"label": "Establecimientos con internación", "unidad": "establecimientos", "agg": "sum"},
    "tasa_mortalidad_hospitalaria": {"label": "Tasa de mortalidad hospitalaria", "unidad": "%", "agg": "mean"},
}
_SALUD_DESAG_ORDEN = [
    "Total", "1 año y más", "De 1 a 4 años", "Menores de 1 año",
    "Neonatal (< 28 días)", "Posneonatal (28 días y más)",
    "Menos de 2.500 g", "2.500 g y más", "Peso sin especificar",
]


def salud():
    fields = ["anio", "departamento", "area_operativa", "desagregacion", "metrica", "valor"]
    rows, notas = [], []

    def _num(df):
        df["anio"] = pd.to_numeric(df["anio"], errors="coerce").astype("Int64")
        df["valor"] = pd.to_numeric(df["valor"], errors="coerce")
        n = int(df["valor"].isna().sum())
        if n:
            notas.append("salud: %d valores no numéricos descartados." % n)
        return df.dropna(subset=["anio", "valor"])

    # -- hechos vitales provinciales (2020-2025) ----------------------------
    hv = _num(pd.read_csv(_drv("Salud_Hechos_Vitales_Salta_provincia_2020-2025.csv"),
                          encoding="utf-8-sig"))
    for r in hv.itertuples(index=False):
        rows.append([int(r.anio), "Salta", "Total", r.desagregacion, r.metrica, float(r.valor)])

    # -- servicios provinciales: consultas e internación (2021-2025) --------
    sv = _num(pd.read_csv(_drv("Salud_Servicios_Salta_provincia_2021-2025.csv"),
                          encoding="utf-8-sig"))
    for r in sv.itertuples(index=False):
        rows.append([int(r.anio), "Salta", "Total", "Total", r.metrica, float(r.valor)])

    # -- nacidos vivos por Área Operativa de residencia (2021-2025) --------
    ao = _num(pd.read_csv(_drv("Salud_Nacidos_Vivos_Salta_area_operativa_2021-2025.csv"),
                          encoding="utf-8-sig"))
    ao["departamento"] = ao["departamento"].map(norm_dept)
    for r in ao.itertuples(index=False):
        rows.append([int(r.anio), r.departamento, r.area_operativa, "Total",
                     r.metrica, float(r.valor)])

    # -- internación por departamento (sólo 2025) --------------------------
    # No trae fila provincial: esa serie ya viene del CSV de servicios y duplicarla haría que
    # los gráficos por departamento contaran dos veces el total.
    it = _num(pd.read_csv(_drv("Salud_Internacion_Salta_departamento_2025.csv"),
                          encoding="utf-8-sig"))
    it["departamento"] = it["departamento"].map(norm_dept)
    anios_it = sorted({int(a) for a in it["anio"]})
    for r in it.itertuples(index=False):
        rows.append([int(r.anio), r.departamento, "Total", "Total", r.metrica, float(r.valor)])

    deptos = [d for d in sorted({r[1] for r in rows})
              if d not in NON_GEO and d not in PROVINCIAL_TOKENS]
    aos = [a for a in sorted({r[2] for r in rows}) if a != "Total"]
    desag = ([d for d in _SALUD_DESAG_ORDEN if d in {r[3] for r in rows}]
             + sorted({r[3] for r in rows} - set(_SALUD_DESAG_ORDEN)))

    notas += [
        "salud: la apertura territorial de nacimientos es por ÁREA OPERATIVA de residencia "
        "(circunscripción sanitaria), no por departamento; el pasaje a departamento lo hace el "
        "script de datos-drive/Actualizar_Salud_Estadisticas_Vitales_Salta.md y es una "
        "aproximación. La suma de las 47 AO coincide exactamente con el total provincial en los "
        "cinco años.",
        "salud: el detalle departamental de internación existe sólo para %s; la serie 2021–2025 "
        "de internación es provincial. Los cocientes por departamento (ocupación, permanencia, "
        "giro, mortalidad) se recalculan sobre los totales sumados de sus establecimientos, no se "
        "promedian." % ", ".join(str(a) for a in anios_it),
        "salud: las tres tablas provinciales del documento fuente se contradicen entre sí "
        "(defunciones 2021/2024/2025, nacidos vivos 2025). Se publica la Tabla 1 (Resumen "
        "Quinquenal), la única internamente consistente y la única que cierra contra la apertura "
        "territorial. El detalle está en datos-drive/Metodologia_Salud_Estadisticas_Vitales_Salta.md.",
        "salud: las tasas de mortalidad general de 2021 y 2022 no se reproducen con la población "
        "que publica el propio documento (la fuente usó proyecciones anteriores a las INDEC "
        "2022–2040); se publican tal como las emitió el MSP.",
        "salud: 2025 es provisorio en toda la fuente.",
    ]
    return {
        "fields": fields,
        "rows": rows,
        "dims": {
            "anio": sorted({r[0] for r in rows}),
            "departamento": deptos,
            "area_operativa": aos,
            "desagregacion": desag,
        },
        "metricas": SALUD_METRICAS,
        "notas": notas,
    }


# Registro id_tema -> función adapter
ADAPTERS = {
    "educacion": educacion,
    "vitivinicultura": inv,
    "produccion-energia": produccion_energia,
    "empleo": empleo,
    "turismo": turismo,
    "agricultura": agricultura,
    "gobierno": gobierno,
    "ganaderia": ganaderia,
    "mineria": mineria,
    "financiero": financiero,
    "construccion": construccion,
    "recursos-municipios": recursos_municipios,
    "energia-electrica": energia_renovable,   # mismo dataset CAMMESA; el catálogo separa qué muestra
    "energia-renovable": energia_renovable,
    "id-innovacion": id_innovacion,
    "resultado-fiscal": resultado_fiscal,
    "recaudacion": recaudacion,
    "exportaciones": exportaciones,
    "salud": salud,
}
